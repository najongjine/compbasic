from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE = Path(__file__).resolve().parent
DOC_DIR = BASE / "항목별교재"
EX_DIR = BASE / "항목별실습예제"
FINAL_DIR = BASE / "최종실습"

HEADER_FILL = "D9EAF7"
NOTE_FILL = "FFF2CC"
GOOD_FILL = "E2F0D9"
WARN_FILL = "FCE4D6"
BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


ITEMS = [
    {
        "no": "01",
        "key": "TODAY",
        "title": "TODAY 함수",
        "formula": "=TODAY()",
        "meaning": "컴퓨터 기준의 오늘 날짜를 자동으로 표시합니다.",
        "scenario": "거래처별 납부 마감일까지 며칠 남았는지 확인합니다.",
        "syntax": "TODAY()",
        "point": "괄호 안에 아무것도 넣지 않습니다. 파일을 다시 열면 날짜가 오늘 날짜로 바뀝니다.",
        "steps": ["B2에 =TODAY()를 입력합니다.", "마감일에서 오늘 날짜를 빼서 남은 일수를 구합니다.", "남은 일수가 음수이면 이미 지난 마감일입니다."],
        "practice": "거래처 미수금 관리표",
    },
    {
        "no": "02",
        "key": "DATE",
        "title": "DATE 함수",
        "formula": "=DATE(2026,6,30)",
        "meaning": "연도, 월, 일을 따로 입력해 정확한 날짜로 합칩니다.",
        "scenario": "계약 시작 연도, 월, 일을 따로 받은 뒤 계약 시작일을 만듭니다.",
        "syntax": "DATE(년, 월, 일)",
        "point": "날짜처럼 보여도 문자로 입력된 값은 계산이 꼬일 수 있습니다. DATE를 쓰면 엑셀이 날짜로 인식합니다.",
        "steps": ["연도, 월, 일을 각각 입력합니다.", "DATE 함수로 하나의 날짜를 만듭니다.", "만든 날짜에 기간을 더해 종료일을 계산합니다."],
        "practice": "계약 기간 관리표",
    },
    {
        "no": "03",
        "key": "YEAR",
        "title": "YEAR 함수",
        "formula": "=YEAR(A2)",
        "meaning": "날짜에서 연도만 꺼냅니다.",
        "scenario": "구입일을 기준으로 비품 구입 연도를 분류합니다.",
        "syntax": "YEAR(날짜)",
        "point": "연도별로 정렬하거나 필터링할 때 날짜 전체보다 연도 열을 따로 만들면 편합니다.",
        "steps": ["구입일 옆에 구입연도 열을 만듭니다.", "YEAR 함수로 연도만 표시합니다.", "연도별로 비품을 확인합니다."],
        "practice": "비품 구입연도 정리표",
    },
    {
        "no": "04",
        "key": "WORKDAY",
        "title": "WORKDAY 함수",
        "formula": "=WORKDAY(A2,5)",
        "meaning": "토요일과 일요일을 빼고 며칠 뒤 근무일을 계산합니다.",
        "scenario": "접수일로부터 5근무일 뒤 처리 예정일을 계산합니다.",
        "syntax": "WORKDAY(시작날짜, 근무일수)",
        "point": "달력 날짜 5일 뒤가 아니라 실제 일하는 날 기준 5일 뒤를 계산합니다.",
        "steps": ["접수일과 처리 근무일수를 입력합니다.", "WORKDAY로 처리 예정일을 계산합니다.", "주말이 끼어도 자동으로 평일 날짜가 나오는지 확인합니다."],
        "practice": "민원/수리 접수 처리일표",
    },
    {
        "no": "05",
        "key": "SUM",
        "title": "SUM 함수",
        "formula": "=SUM(B2:B10)",
        "meaning": "선택한 숫자를 모두 더합니다.",
        "scenario": "일자별 매출을 더해 주간 매출 합계를 구합니다.",
        "syntax": "SUM(더할 범위)",
        "point": "숫자가 들어 있는 칸만 더합니다. 문자나 빈칸은 합계에서 제외됩니다.",
        "steps": ["매출액 범위를 드래그합니다.", "SUM 함수로 합계를 구합니다.", "중간에 숫자를 바꾸면 합계가 자동으로 바뀌는지 확인합니다."],
        "practice": "주간 매출 합계표",
    },
    {
        "no": "06",
        "key": "COUNT_COUNTA",
        "title": "COUNT와 COUNTA 함수",
        "formula": "=COUNT(C2:C10), =COUNTA(A2:A10)",
        "meaning": "COUNT는 숫자 칸 개수, COUNTA는 비어 있지 않은 칸 개수를 셉니다.",
        "scenario": "상담 기록에서 금액이 입력된 건수와 전체 상담 건수를 비교합니다.",
        "syntax": "COUNT(범위), COUNTA(범위)",
        "point": "COUNT는 숫자만 셉니다. 이름, 메모 같은 글자는 COUNTA로 셉니다.",
        "steps": ["상담명 범위에 COUNTA를 사용합니다.", "견적금액 범위에 COUNT를 사용합니다.", "상담은 했지만 견적이 없는 건수를 비교합니다."],
        "practice": "상담 기록 입력현황표",
    },
    {
        "no": "07",
        "key": "MAX",
        "title": "MAX 함수",
        "formula": "=MAX(B2:B10)",
        "meaning": "범위 안에서 가장 큰 숫자를 찾습니다.",
        "scenario": "상품별 매출 중 가장 높은 매출액을 찾습니다.",
        "syntax": "MAX(범위)",
        "point": "어느 값이 가장 큰지만 빠르게 확인할 때 좋습니다.",
        "steps": ["매출액 범위를 선택합니다.", "MAX 함수로 최고 매출액을 구합니다.", "가장 큰 값이 어느 품목인지 표에서 확인합니다."],
        "practice": "품목별 최고 매출 확인표",
    },
    {
        "no": "08",
        "key": "MIN",
        "title": "MIN 함수",
        "formula": "=MIN(B2:B10)",
        "meaning": "범위 안에서 가장 작은 숫자를 찾습니다.",
        "scenario": "지점별 재고 중 가장 적은 재고 수량을 찾습니다.",
        "syntax": "MIN(범위)",
        "point": "재고, 잔액, 점검 수치처럼 낮을수록 확인이 필요한 자료에 자주 씁니다.",
        "steps": ["재고수량 범위를 선택합니다.", "MIN 함수로 최저 재고를 구합니다.", "최저 재고 품목을 확인합니다."],
        "practice": "지점별 재고 최저값 확인표",
    },
    {
        "no": "09",
        "key": "IF",
        "title": "IF 함수",
        "formula": '=IF(C2>=D2,"정상","확인")',
        "meaning": "조건이 맞으면 첫 번째 결과, 아니면 두 번째 결과를 표시합니다.",
        "scenario": "입금액이 청구액 이상이면 완료, 부족하면 확인으로 표시합니다.",
        "syntax": "IF(조건, 맞을 때 값, 아닐 때 값)",
        "point": "IF는 엑셀에서 가장 자주 쓰는 판단 함수입니다. 말로는 '만약 ~라면'입니다.",
        "steps": ["청구액과 입금액을 비교합니다.", "입금액이 청구액 이상이면 완료로 표시합니다.", "아니면 확인으로 표시합니다."],
        "practice": "거래처 입금 확인표",
    },
    {
        "no": "10",
        "key": "IF_AND",
        "title": "IF와 AND 함수",
        "formula": '=IF(AND(C2>=D2,E2<=2%),"출고가능","확인")',
        "meaning": "조건이 모두 맞을 때만 원하는 결과를 표시합니다.",
        "scenario": "수량이 충분하고 불량률이 낮을 때만 출고가능으로 표시합니다.",
        "syntax": "IF(AND(조건1, 조건2), 맞을 때 값, 아닐 때 값)",
        "point": "AND는 '그리고'입니다. 조건이 하나라도 틀리면 전체가 틀린 것으로 봅니다.",
        "steps": ["납품수량이 주문수량 이상인지 봅니다.", "불량률이 기준 이하인지 봅니다.", "두 조건이 모두 맞을 때 출고가능으로 표시합니다."],
        "practice": "납품 품질검수표",
    },
    {
        "no": "11",
        "key": "IF_OR",
        "title": "IF와 OR 함수",
        "formula": '=IF(OR(C2<D2,E2<=7),"긴급확인","정상")',
        "meaning": "조건 중 하나라도 맞으면 원하는 결과를 표시합니다.",
        "scenario": "재고가 부족하거나 유통기한이 임박하면 긴급확인으로 표시합니다.",
        "syntax": "IF(OR(조건1, 조건2), 맞을 때 값, 아닐 때 값)",
        "point": "OR는 '또는'입니다. 여러 위험 조건 중 하나만 걸려도 표시할 때 씁니다.",
        "steps": ["현재 재고가 안전재고보다 적은지 확인합니다.", "유통기한까지 남은 일수가 7일 이하인지 확인합니다.", "둘 중 하나라도 맞으면 긴급확인으로 표시합니다."],
        "practice": "재고 긴급점검표",
    },
    {
        "no": "12",
        "key": "COUNTIF",
        "title": "COUNTIF 함수",
        "formula": '=COUNTIF(C2:C20,"카드")',
        "meaning": "조건에 맞는 칸의 개수를 셉니다.",
        "scenario": "결제수단별 거래 건수를 셉니다.",
        "syntax": "COUNTIF(조건을 볼 범위, 조건)",
        "point": "몇 건인지 셀 때 씁니다. 금액 합계가 필요하면 SUMIF를 씁니다.",
        "steps": ["결제수단 범위를 선택합니다.", "카드, 현금, 계좌이체 조건별로 COUNTIF를 씁니다.", "어떤 결제수단이 가장 많은지 확인합니다."],
        "practice": "결제수단별 거래건수표",
    },
    {
        "no": "13",
        "key": "AVERAGE",
        "title": "AVERAGE 함수",
        "formula": "=AVERAGE(B2:D2)",
        "meaning": "숫자의 평균을 구합니다.",
        "scenario": "최근 3개월 매출 평균을 구해 평소 수준을 확인합니다.",
        "syntax": "AVERAGE(범위)",
        "point": "합계를 개수로 나누는 계산을 함수가 대신합니다.",
        "steps": ["3개월 매출 범위를 선택합니다.", "AVERAGE 함수로 평균을 구합니다.", "이번 달 매출이 평균보다 높은지 비교합니다."],
        "practice": "3개월 매출 평균표",
    },
    {
        "no": "14",
        "key": "RANK_EQ",
        "title": "RANK.EQ 함수",
        "formula": "=RANK.EQ(B2,$B$2:$B$10,0)",
        "meaning": "값의 순위를 매깁니다. 0을 넣으면 큰 값이 1등입니다.",
        "scenario": "상품별 매출 순위를 계산합니다.",
        "syntax": "RANK.EQ(순위를 매길 값, 전체 범위, 0)",
        "point": "$B$2:$B$10처럼 달러 표시를 붙이면 수식을 복사해도 전체 범위가 고정됩니다.",
        "steps": ["첫 번째 상품 매출의 순위를 구합니다.", "전체 매출 범위는 F4로 고정합니다.", "수식을 아래로 복사해 모든 상품 순위를 구합니다."],
        "practice": "상품 매출 순위표",
    },
    {
        "no": "15",
        "key": "SUMIF",
        "title": "SUMIF 함수",
        "formula": '=SUMIF(B2:B20,"식자재",E2:E20)',
        "meaning": "조건에 맞는 행의 금액만 더합니다.",
        "scenario": "비용 내역에서 항목별 지출 합계를 구합니다.",
        "syntax": "SUMIF(조건범위, 조건, 합계범위)",
        "point": "조건을 볼 범위와 더할 범위가 다를 수 있습니다. 범위 크기는 같게 잡습니다.",
        "steps": ["비용분류 범위를 조건범위로 잡습니다.", "합계를 낼 금액 범위를 잡습니다.", "식자재, 임대료, 공과금별 합계를 계산합니다."],
        "practice": "비용 항목별 합계표",
    },
    {
        "no": "16",
        "key": "SORT",
        "title": "정렬 기능",
        "formula": "데이터 > 정렬",
        "meaning": "표 전체를 원하는 기준 순서대로 다시 배치합니다.",
        "scenario": "미수금이 큰 거래처부터 확인합니다.",
        "syntax": "표 안의 셀 선택 > 데이터 > 정렬 > 기준 선택",
        "point": "한 열만 드래그해서 정렬하면 다른 열과 행이 어긋날 수 있습니다. 표 안의 아무 셀 하나만 선택하고 정렬합니다.",
        "steps": ["표 안의 아무 셀 하나를 클릭합니다.", "데이터 탭에서 정렬을 누릅니다.", "미수금 기준, 큰 값에서 작은 값 순으로 정렬합니다."],
        "practice": "거래처 미수금 정렬표",
    },
    {
        "no": "17",
        "key": "FILTER",
        "title": "필터 기능",
        "formula": "데이터 > 필터",
        "meaning": "전체 자료 중 필요한 조건의 행만 잠시 보여줍니다.",
        "scenario": "입금확인이 필요한 거래처만 봅니다.",
        "syntax": "표 안의 셀 선택 > 데이터 > 필터",
        "point": "필터는 데이터를 지우는 기능이 아닙니다. 조건에 맞지 않는 행을 잠시 숨기는 기능입니다.",
        "steps": ["표 안의 아무 셀 하나를 클릭합니다.", "데이터 탭에서 필터를 켭니다.", "상태 열에서 입금확인만 선택합니다."],
        "practice": "거래처 상태 필터표",
    },
    {
        "no": "18",
        "key": "CONDITIONAL_FORMATTING",
        "title": "조건부 서식",
        "formula": "홈 > 조건부 서식",
        "meaning": "조건에 맞는 셀을 자동으로 색칠합니다.",
        "scenario": "재고가 안전재고보다 적은 품목을 자동으로 표시합니다.",
        "syntax": "홈 > 조건부 서식 > 셀 강조 규칙 또는 새 규칙",
        "point": "눈으로 찾기 힘든 위험 자료를 자동으로 표시할 때 좋습니다.",
        "steps": ["현재재고 범위를 선택합니다.", "조건부 서식에서 기준보다 작은 값을 선택합니다.", "안전재고보다 부족한 품목을 색으로 확인합니다."],
        "practice": "재고 부족 자동표시표",
    },
]


def clean_dirs() -> None:
    for directory in [DOC_DIR, EX_DIR, FINAL_DIR]:
        directory.mkdir(exist_ok=True)
        for path in directory.glob("*"):
            if path.is_file():
                path.unlink()


def write_md(item: dict[str, str]) -> None:
    text = f"""# {item['title']}

## 한 줄 설명
{item['meaning']}

## 실제로 쓰는 상황
{item['scenario']}

## 기본 모양
`{item['syntax']}`

## 예시
`{item['formula']}`

## 쉽게 설명하면
엑셀 함수는 어려운 영어 문장이 아니라, 반복 계산을 대신 해주는 짧은 명령입니다.
이 항목에서는 `{item['title']}`을(를) 이용해 `{item['practice']}`를 처리합니다.

## 수업에서 꼭 짚을 점
{item['point']}

## 실습 순서
"""
    for idx, step in enumerate(item["steps"], start=1):
        text += f"{idx}. {step}\n"
    text += f"""
## 학생에게 물어볼 확인 질문
- 이 기능을 쓰면 손으로 계산할 때보다 무엇이 편해지나요?
- 함수 괄호 안에는 어떤 값을 넣었나요?
- 결과가 이상하면 먼저 어느 범위를 확인해야 하나요?

## 연결 실습 파일
`{item['no']}_{item['key']}_{item['practice']}.xlsx`
"""
    (DOC_DIR / f"{item['no']}_{item['key']}_{item['title']}.md").write_text(text, encoding="utf-8")


def style_sheet(ws, title: str, subtitle: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = title
    ws["A1"].font = Font(size=15, bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(color="666666")
    ws.row_dimensions[1].height = 24


def style_table(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
    for cell in ws[min_row][min_col - 1 : max_col]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_excel_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def set_widths(ws, widths: list[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def money(ws, cells: list[str]) -> None:
    for cell in cells:
        ws[cell].number_format = '#,##0"원"'


def save(wb: Workbook, path: Path) -> None:
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    wb.save(path)


def safe_name(text: str) -> str:
    for old in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
        text = text.replace(old, "_")
    return text


def add_instruction_sheet(wb: Workbook, item: dict[str, str]) -> None:
    ws = wb.create_sheet("수업진행")
    style_sheet(ws, item["title"], item["meaning"], 3)
    ws.append([])
    ws.append(["순서", "강사가 할 일", "수강생이 할 일"])
    for idx, step in enumerate(item["steps"], start=1):
        ws.append([idx, step, "같이 입력하고 결과 확인"])
    style_table(ws, 4, 3 + len(item["steps"]), 1, 3)
    set_widths(ws, [8, 42, 24])


def workbook_for_item(item: dict[str, str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "실습"
    no = item["no"]
    key = item["key"]
    style_sheet(ws, f"{item['title']} 실습", item["scenario"], 8)
    ws.append([])

    if key == "TODAY":
        ws["A4"] = "오늘 날짜"
        ws["B4"] = "=TODAY()"
        ws["B4"].fill = PatternFill("solid", fgColor=NOTE_FILL)
        ws.append(["거래처", "청구액", "마감일", "남은일수", "상태"])
        data = [
            ["한빛상사", 280000, date.today() + timedelta(days=5)],
            ["서울푸드", 430000, date.today() - timedelta(days=2)],
            ["미래유통", 165000, date.today() + timedelta(days=12)],
            ["도담상회", 92000, date.today()],
        ]
        for r, row in enumerate(data, start=6):
            ws.append(row + [f"=C{r}-$B$4", f'=IF(D{r}<0,"기한지남","확인")'])
        style_table(ws, 5, 9, 1, 5)
        add_excel_table(ws, "tblToday", "A5:E9")
        set_widths(ws, [14, 12, 12, 10, 12])
        for r in range(6, 10):
            ws[f"B{r}"].number_format = '#,##0"원"'
            ws[f"C{r}"].number_format = "yyyy-mm-dd"
        ws["B4"].number_format = "yyyy-mm-dd"
    elif key == "DATE":
        ws.append(["계약명", "연도", "월", "일", "계약시작일", "계약개월", "계약종료일"])
        rows = [["청소용역", 2026, 6, 1, 12], ["복합기임대", 2026, 7, 15, 24], ["식자재공급", 2026, 8, 1, 6], ["보안관리", 2026, 9, 10, 12]]
        for r, row in enumerate(rows, start=5):
            ws.append(row[:4] + [f"=DATE(B{r},C{r},D{r})", row[4], f"=E{r}+F{r}*30-1"])
        style_table(ws, 4, 8, 1, 7)
        add_excel_table(ws, "tblDate", "A4:G8")
        set_widths(ws, [14, 8, 8, 8, 13, 10, 13])
        for r in range(5, 9):
            ws[f"E{r}"].number_format = "yyyy-mm-dd"
            ws[f"G{r}"].number_format = "yyyy-mm-dd"
    elif key == "YEAR":
        ws.append(["비품명", "구입일", "구입연도", "구입금액", "관리부서"])
        rows = [["노트북", date(2024, 3, 12), 980000, "사무"], ["프린터", date(2025, 1, 20), 310000, "총무"], ["냉장고", date(2023, 8, 5), 560000, "매장"], ["카드단말기", date(2026, 2, 1), 180000, "매장"]]
        for r, row in enumerate(rows, start=5):
            ws.append([row[0], row[1], f"=YEAR(B{r})", row[2], row[3]])
        style_table(ws, 4, 8, 1, 5)
        add_excel_table(ws, "tblYear", "A4:E8")
        set_widths(ws, [14, 12, 10, 12, 10])
        for r in range(5, 9):
            ws[f"B{r}"].number_format = "yyyy-mm-dd"
            ws[f"D{r}"].number_format = '#,##0"원"'
    elif key == "WORKDAY":
        ws.append(["접수번호", "업무내용", "접수일", "처리근무일", "처리예정일"])
        rows = [["R-001", "간판 수리", date(2026, 6, 5), 5], ["R-002", "거래명세서 재발행", date(2026, 6, 8), 2], ["R-003", "설비 점검", date(2026, 6, 12), 4], ["R-004", "환불 검토", date(2026, 6, 15), 3]]
        for r, row in enumerate(rows, start=5):
            ws.append(row + [f"=WORKDAY(C{r},D{r})"])
        style_table(ws, 4, 8, 1, 5)
        add_excel_table(ws, "tblWorkday", "A4:E8")
        set_widths(ws, [12, 20, 12, 12, 13])
        for r in range(5, 9):
            ws[f"C{r}"].number_format = "yyyy-mm-dd"
            ws[f"E{r}"].number_format = "yyyy-mm-dd"
    elif key == "SUM":
        ws.append(["일자", "현금매출", "카드매출", "배달매출", "일매출"])
        for r, row in enumerate([[1, 125000, 340000, 86000], [2, 98000, 410000, 76000], [3, 142000, 386000, 95000], [4, 87000, 295000, 64000], [5, 156000, 430000, 110000]], start=5):
            ws.append([f"6월 {row[0]}일"] + row[1:] + [f"=SUM(B{r}:D{r})"])
        ws.append(["주간합계", "=SUM(B5:B9)", "=SUM(C5:C9)", "=SUM(D5:D9)", "=SUM(E5:E9)"])
        style_table(ws, 4, 10, 1, 5)
        set_widths(ws, [12, 12, 12, 12, 12])
        for row in range(5, 11):
            for col in "BCDE":
                ws[f"{col}{row}"].number_format = '#,##0"원"'
    elif key == "COUNT_COUNTA":
        ws.append(["상담일", "고객명", "상담내용", "견적금액", "계약여부"])
        rows = [[date(2026, 6, 1), "김민수", "간판 교체", 480000, "진행"], [date(2026, 6, 2), "박영희", "메뉴판 제작", "", "대기"], [date(2026, 6, 3), "이정훈", "복합기 임대", 120000, "완료"], [date(2026, 6, 4), "최은정", "청소용역 문의", "", "대기"], [date(2026, 6, 5), "정상호", "비품 납품", 260000, "진행"]]
        for row in rows:
            ws.append(row)
        ws["G4"], ws["H4"] = "항목", "결과"
        ws["G5"], ws["H5"] = "전체 상담 건수", "=COUNTA(B5:B9)"
        ws["G6"], ws["H6"] = "견적 입력 건수", "=COUNT(D5:D9)"
        style_table(ws, 4, 9, 1, 5)
        style_table(ws, 4, 6, 7, 8)
        set_widths(ws, [12, 12, 18, 12, 10, 4, 16, 10])
        for r in range(5, 10):
            ws[f"A{r}"].number_format = "yyyy-mm-dd"
            ws[f"D{r}"].number_format = '#,##0"원"'
    elif key in {"MAX", "MIN", "RANK_EQ"}:
        ws.append(["품목", "분류", "판매수량", "매출액", "순위" if key == "RANK_EQ" else "비고"])
        rows = [["아메리카노", "음료", 138, 441600], ["카페라떼", "음료", 94, 394800], ["샌드위치", "간편식", 70, 455000], ["샐러드", "간편식", 45, 324000], ["원두 200g", "상품", 28, 336000]]
        for r, row in enumerate(rows, start=5):
            extra = f"=RANK.EQ(D{r},$D$5:$D$9,0)" if key == "RANK_EQ" else ""
            ws.append(row + [extra])
        ws["G4"], ws["H4"] = "계산", "결과"
        if key == "MAX":
            ws["G5"], ws["H5"] = "최고 매출액", "=MAX(D5:D9)"
        elif key == "MIN":
            ws["G5"], ws["H5"] = "최저 매출액", "=MIN(D5:D9)"
        else:
            ws["G5"], ws["H5"] = "1위 확인", "순위 열에서 1 확인"
        style_table(ws, 4, 9, 1, 5)
        style_table(ws, 4, 5, 7, 8)
        set_widths(ws, [14, 10, 10, 12, 10, 4, 14, 14])
        for r in range(5, 10):
            ws[f"D{r}"].number_format = '#,##0"원"'
        ws["H5"].number_format = '#,##0"원"'
    elif key == "IF":
        ws.append(["거래처", "청구액", "입금액", "미수금", "상태"])
        rows = [["서울푸드", 430000, 430000], ["한빛문구", 230000, 180000], ["미래유통", 520000, 300000], ["도담상회", 98000, 98000]]
        for r, row in enumerate(rows, start=5):
            ws.append(row + [f"=B{r}-C{r}", f'=IF(C{r}>=B{r},"완료","확인")'])
        style_table(ws, 4, 8, 1, 5)
        set_widths(ws, [14, 12, 12, 12, 10])
        for r in range(5, 9):
            for col in "BCD":
                ws[f"{col}{r}"].number_format = '#,##0"원"'
    elif key == "IF_AND":
        ws.append(["납품처", "주문수량", "납품수량", "불량률", "판정"])
        rows = [["서울푸드", 100, 100, 0.01], ["한빛문구", 80, 76, 0.01], ["미래유통", 120, 120, 0.035], ["도담상회", 60, 62, 0.0]]
        for r, row in enumerate(rows, start=5):
            ws.append(row + [f'=IF(AND(C{r}>=B{r},D{r}<=2%),"출고가능","확인")'])
        style_table(ws, 4, 8, 1, 5)
        set_widths(ws, [14, 10, 10, 10, 12])
        for r in range(5, 9):
            ws[f"D{r}"].number_format = "0.0%"
    elif key == "IF_OR":
        ws.append(["품목", "현재재고", "안전재고", "유통기한까지", "상태"])
        rows = [["우유", 14, 12, 5], ["샌드위치 포장지", 45, 50, 40], ["샐러드 용기", 30, 25, 20], ["원두 200g", 6, 10, 35]]
        for r, row in enumerate(rows, start=5):
            ws.append(row + [f'=IF(OR(B{r}<C{r},D{r}<=7),"긴급확인","정상")'])
        style_table(ws, 4, 8, 1, 5)
        set_widths(ws, [18, 10, 10, 12, 12])
    elif key == "COUNTIF":
        ws.append(["일자", "거래처", "결제수단", "금액"])
        rows = [[date(2026, 6, 1), "서울푸드", "카드", 120000], [date(2026, 6, 1), "한빛문구", "현금", 80000], [date(2026, 6, 2), "미래유통", "카드", 160000], [date(2026, 6, 2), "도담상회", "계좌이체", 90000], [date(2026, 6, 3), "서울푸드", "카드", 110000]]
        for row in rows:
            ws.append(row)
        ws["F4"], ws["G4"] = "결제수단", "거래건수"
        for r, method in enumerate(["카드", "현금", "계좌이체"], start=5):
            ws[f"F{r}"] = method
            ws[f"G{r}"] = f"=COUNTIF(C5:C9,F{r})"
        style_table(ws, 4, 9, 1, 4)
        style_table(ws, 4, 7, 6, 7)
        set_widths(ws, [12, 14, 12, 12, 4, 12, 10])
        for r in range(5, 10):
            ws[f"A{r}"].number_format = "yyyy-mm-dd"
            ws[f"D{r}"].number_format = '#,##0"원"'
    elif key == "AVERAGE":
        ws.append(["품목", "4월", "5월", "6월", "3개월 평균", "판정"])
        rows = [["아메리카노", 390000, 420000, 450000], ["카페라떼", 310000, 360000, 390000], ["샌드위치", 450000, 430000, 470000], ["샐러드", 280000, 300000, 330000]]
        for r, row in enumerate(rows, start=5):
            ws.append(row + [f"=AVERAGE(B{r}:D{r})", f'=IF(D{r}>=E{r},"평균이상","평균미만")'])
        style_table(ws, 4, 8, 1, 6)
        set_widths(ws, [14, 12, 12, 12, 12, 12])
        for r in range(5, 9):
            for col in "BCDE":
                ws[f"{col}{r}"].number_format = '#,##0"원"'
    elif key == "SUMIF":
        ws.append(["일자", "비용분류", "내용", "금액"])
        rows = [[date(2026, 6, 1), "식자재", "원두/우유", 185000], [date(2026, 6, 2), "운영비", "소모품", 76000], [date(2026, 6, 5), "고정비", "임대료", 850000], [date(2026, 6, 7), "식자재", "빵/야채", 132000], [date(2026, 6, 9), "운영비", "전기요금", 118000]]
        for row in rows:
            ws.append(row)
        ws["F4"], ws["G4"] = "비용분류", "합계"
        for r, cat in enumerate(["식자재", "운영비", "고정비"], start=5):
            ws[f"F{r}"] = cat
            ws[f"G{r}"] = f"=SUMIF(B5:B9,F{r},D5:D9)"
        style_table(ws, 4, 9, 1, 4)
        style_table(ws, 4, 7, 6, 7)
        set_widths(ws, [12, 12, 18, 12, 4, 12, 12])
        for r in range(5, 10):
            ws[f"A{r}"].number_format = "yyyy-mm-dd"
            ws[f"D{r}"].number_format = '#,##0"원"'
        for r in range(5, 8):
            ws[f"G{r}"].number_format = '#,##0"원"'
    elif key in {"SORT", "FILTER"}:
        ws.append(["거래처", "지역", "청구액", "입금액", "미수금", "납부예정일", "상태"])
        rows = [["서울푸드", "서울", 430000, 430000, date(2026, 6, 10)], ["한빛문구", "경기", 230000, 180000, date(2026, 6, 12)], ["미래유통", "서울", 520000, 300000, date(2026, 6, 15)], ["도담상회", "인천", 98000, 98000, date(2026, 6, 8)], ["청솔물류", "충남", 610000, 450000, date(2026, 6, 18)]]
        for r, row in enumerate(rows, start=5):
            ws.append(row[:4] + [f"=C{r}-D{r}", row[4], f'=IF(E{r}>0,"입금확인","완료")'])
        style_table(ws, 4, 9, 1, 7)
        add_excel_table(ws, f"tbl{key}", "A4:G9")
        ws.auto_filter.ref = "A4:G9"
        set_widths(ws, [14, 10, 12, 12, 12, 13, 12])
        for r in range(5, 10):
            for col in "CDE":
                ws[f"{col}{r}"].number_format = '#,##0"원"'
            ws[f"F{r}"].number_format = "yyyy-mm-dd"
    elif key == "CONDITIONAL_FORMATTING":
        ws.append(["품목", "현재재고", "안전재고", "차이", "상태"])
        rows = [["아메리카노 원두", 8, 10], ["우유", 14, 12], ["샌드위치 포장지", 45, 50], ["샐러드 용기", 30, 25], ["원두 200g", 6, 10]]
        for r, row in enumerate(rows, start=5):
            ws.append(row + [f"=B{r}-C{r}", f'=IF(B{r}<C{r},"보충필요","정상")'])
        style_table(ws, 4, 9, 1, 5)
        ws.conditional_formatting.add("E5:E9", CellIsRule(operator="equal", formula=['"보충필요"'], fill=PatternFill("solid", fgColor=WARN_FILL)))
        ws.conditional_formatting.add("D5:D9", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=WARN_FILL)))
        set_widths(ws, [18, 10, 10, 8, 12])

    ws.freeze_panes = "A5"
    add_instruction_sheet(wb, item)
    save(wb, EX_DIR / safe_name(f"{item['no']}_{item['key']}_{item['practice']}.xlsx"))


def create_final_project() -> None:
    md = """# 최종 실습 프로젝트: 소상공인 월간 매출정산표

## 실습 목표
작은 매장이나 개인사업자가 자주 쓰는 월간 매출, 비용, 미수금, 재고 확인표를 완성합니다.

## 억지로 넣지 않은 것
모든 함수를 억지로 넣지 않습니다. 실제로 자주 쓰는 합계, 조건별 합계, 건수, 평균, IF 판단, 정렬, 필터, 조건부 서식만 사용합니다.

## 완성 후 할 수 있어야 하는 말
- 이번 달 총매출과 총비용은 얼마인가요?
- 순이익은 얼마인가요?
- 결제수단별 매출은 어떻게 나뉘나요?
- 미수금이 남은 거래처는 어디인가요?
- 재고를 먼저 보충해야 할 품목은 무엇인가요?

## 사용 파일
`최종실습_소상공인_월간매출정산.xlsx`
"""
    (FINAL_DIR / "최종실습_소상공인_월간매출정산.md").write_text(md, encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"
    style_sheet(ws, "소상공인 월간 매출정산", "매출과 비용을 한 달 단위로 정리합니다.", 10)
    ws.append([])
    ws.append(["일자", "구분", "분류", "내용", "결제수단", "수량", "단가", "매출", "비용", "순액"])
    rows = [
        [date(2026, 6, 1), "매출", "음료", "아메리카노", "카드", 42, 3200, 0],
        [date(2026, 6, 1), "매출", "간편식", "샌드위치", "현금", 24, 6500, 0],
        [date(2026, 6, 2), "비용", "식자재", "원두/우유 구입", "계좌이체", 1, 0, 185000],
        [date(2026, 6, 3), "매출", "음료", "카페라떼", "카드", 30, 4200, 0],
        [date(2026, 6, 4), "비용", "운영비", "소모품 구입", "카드", 1, 0, 76000],
        [date(2026, 6, 5), "매출", "상품", "원두 200g", "계좌이체", 12, 12000, 0],
        [date(2026, 6, 8), "비용", "고정비", "임대료", "계좌이체", 1, 0, 850000],
        [date(2026, 6, 9), "매출", "간편식", "샐러드", "카드", 18, 7200, 0],
        [date(2026, 6, 10), "매출", "음료", "아메리카노", "카드", 50, 3200, 0],
        [date(2026, 6, 11), "비용", "고정비", "인건비", "계좌이체", 1, 0, 620000],
    ]
    for r, row in enumerate(rows, start=5):
        ws.append(row[:7] + [f'=IF(B{r}="매출",F{r}*G{r},0)', row[7], f"=H{r}-I{r}"])
    style_table(ws, 4, 14, 1, 10)
    ws.auto_filter.ref = "A4:J14"
    ws.freeze_panes = "A5"
    set_widths(ws, [12, 8, 10, 18, 12, 8, 10, 12, 12, 12])
    for r in range(5, 15):
        ws[f"A{r}"].number_format = "yyyy-mm-dd"
        for col in "GHIJ":
            ws[f"{col}{r}"].number_format = '#,##0"원"'

    summary = wb.create_sheet("요약")
    style_sheet(summary, "월간 요약", "자주 쓰는 함수만 사용한 실무형 요약입니다.", 6)
    summary.append([])
    summary.append(["항목", "값", "사용 기능"])
    summary_rows = [
        ["총매출", "=SUM(거래내역!H5:H14)", "SUM"],
        ["총비용", "=SUM(거래내역!I5:I14)", "SUM"],
        ["순이익", "=B5-B6", "SUM 결과 계산"],
        ["거래건수", "=COUNTA(거래내역!A5:A14)", "COUNTA"],
        ["평균 순액", "=AVERAGE(거래내역!J5:J14)", "AVERAGE"],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.append([])
    summary.append(["결제수단", "매출합계", "거래건수"])
    for r, method in enumerate(["카드", "현금", "계좌이체"], start=12):
        summary.append([method, f"=SUMIF(거래내역!E5:E14,A{r},거래내역!H5:H14)", f"=COUNTIF(거래내역!E5:E14,A{r})"])
    style_table(summary, 4, 9, 1, 3)
    style_table(summary, 11, 14, 1, 3)
    set_widths(summary, [14, 14, 14])
    for r in list(range(5, 10)) + list(range(12, 15)):
        summary[f"B{r}"].number_format = '#,##0"원"'
    clients = wb.create_sheet("미수금관리")
    style_sheet(clients, "미수금 관리", "정렬과 필터로 입금확인 대상을 찾습니다.", 7)
    clients.append([])
    clients.append(["거래처", "청구액", "입금액", "미수금", "납부예정일", "상태", "메모"])
    crows = [["서울푸드", 430000, 430000, date(2026, 6, 10)], ["한빛문구", 230000, 180000, date(2026, 6, 12)], ["미래유통", 520000, 300000, date(2026, 6, 15)], ["도담상회", 98000, 98000, date(2026, 6, 8)]]
    for r, row in enumerate(crows, start=5):
        clients.append(row[:3] + [f"=B{r}-C{r}", row[3], f'=IF(D{r}>0,"입금확인","완료")', ""])
    style_table(clients, 4, 8, 1, 7)
    clients.auto_filter.ref = "A4:G8"
    clients.freeze_panes = "A5"
    clients.conditional_formatting.add("D5:D8", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=WARN_FILL)))
    set_widths(clients, [14, 12, 12, 12, 13, 12, 16])
    for r in range(5, 9):
        for col in "BCD":
            clients[f"{col}{r}"].number_format = '#,##0"원"'
        clients[f"E{r}"].number_format = "yyyy-mm-dd"

    inv = wb.create_sheet("재고관리")
    style_sheet(inv, "재고 관리", "조건부 서식으로 부족한 품목을 확인합니다.", 6)
    inv.append([])
    inv.append(["품목", "현재재고", "안전재고", "차이", "상태", "발주메모"])
    irows = [["아메리카노 원두", 8, 10], ["우유", 14, 12], ["샌드위치 포장지", 45, 50], ["원두 200g", 6, 10]]
    for r, row in enumerate(irows, start=5):
        inv.append(row + [f"=B{r}-C{r}", f'=IF(B{r}<C{r},"보충필요","정상")', ""])
    style_table(inv, 4, 8, 1, 6)
    inv.auto_filter.ref = "A4:F8"
    inv.freeze_panes = "A5"
    inv.conditional_formatting.add("E5:E8", CellIsRule(operator="equal", formula=['"보충필요"'], fill=PatternFill("solid", fgColor=WARN_FILL)))
    set_widths(inv, [18, 10, 10, 8, 12, 16])

    save(wb, FINAL_DIR / "최종실습_소상공인_월간매출정산.xlsx")


def create_household_budget_project() -> None:
    md = """# 최종 실습 프로젝트: 우리집 월간 가계부

## 실습 목표
한 달 수입과 지출을 정리해서 남은 돈, 항목별 지출, 고정비 비중, 예산 초과 항목을 확인합니다.

## 실제로 자주 쓰는 기능
- SUM으로 총수입, 총지출, 잔액을 계산합니다.
- SUMIF로 식비, 교통, 통신비 같은 항목별 지출을 계산합니다.
- COUNTIF로 지출 건수를 셉니다.
- AVERAGE로 하루 평균 지출을 확인합니다.
- IF로 예산 초과 여부를 표시합니다.
- 필터로 특정 항목이나 결제수단만 확인합니다.
- 조건부 서식으로 예산 초과와 큰 지출을 표시합니다.

## 완성 후 확인 질문
- 이번 달 총수입과 총지출은 얼마인가요?
- 가장 많이 쓴 항목은 무엇인가요?
- 예산을 초과한 항목은 무엇인가요?
- 카드 지출과 현금 지출 중 어느 쪽이 더 큰가요?
- 다음 달에 줄일 수 있는 지출은 무엇인가요?

## 사용 파일
`최종실습_우리집_월간가계부.xlsx`
"""
    (FINAL_DIR / "최종실습_우리집_월간가계부.md").write_text(md, encoding="utf-8")

    wb = Workbook()

    ws = wb.active
    ws.title = "가계부"
    style_sheet(ws, "우리집 월간 가계부", "수입과 지출을 한 달 단위로 정리합니다.", 9)
    ws.append([])
    ws.append(["날짜", "구분", "항목", "내용", "결제수단", "수입", "지출", "메모", "확인"])
    rows = [
        [date(2026, 6, 1), "수입", "급여", "월급", "계좌이체", 2600000, 0, ""],
        [date(2026, 6, 1), "지출", "주거", "월세/관리비", "계좌이체", 0, 650000, "고정비"],
        [date(2026, 6, 2), "지출", "식비", "마트 장보기", "카드", 0, 86400, ""],
        [date(2026, 6, 3), "지출", "교통", "교통카드 충전", "카드", 0, 50000, ""],
        [date(2026, 6, 5), "지출", "통신", "휴대폰 요금", "계좌이체", 0, 69000, "고정비"],
        [date(2026, 6, 6), "지출", "식비", "외식", "카드", 0, 42000, ""],
        [date(2026, 6, 8), "지출", "의료", "병원/약국", "카드", 0, 31500, ""],
        [date(2026, 6, 10), "수입", "기타수입", "중고거래", "현금", 70000, 0, ""],
        [date(2026, 6, 12), "지출", "생활용품", "세제/휴지", "카드", 0, 38400, ""],
        [date(2026, 6, 15), "지출", "보험", "보험료", "계좌이체", 0, 145000, "고정비"],
        [date(2026, 6, 18), "지출", "식비", "마트 장보기", "카드", 0, 72300, ""],
        [date(2026, 6, 20), "지출", "문화", "영화/도서", "카드", 0, 28000, ""],
        [date(2026, 6, 22), "지출", "경조사", "축의금", "현금", 0, 100000, ""],
        [date(2026, 6, 25), "지출", "식비", "외식", "카드", 0, 55000, ""],
        [date(2026, 6, 28), "지출", "기타", "예비비 사용", "현금", 0, 30000, ""],
    ]
    for r, row in enumerate(rows, start=5):
        ws.append(row + [f'=IF(G{r}>=100000,"큰지출","")'])
    style_table(ws, 4, 19, 1, 9)
    ws.auto_filter.ref = "A4:I19"
    ws.freeze_panes = "A5"
    set_widths(ws, [12, 8, 12, 18, 12, 12, 12, 12, 10])
    for r in range(5, 20):
        ws[f"A{r}"].number_format = "yyyy-mm-dd"
        ws[f"F{r}"].number_format = '#,##0"원"'
        ws[f"G{r}"].number_format = '#,##0"원"'
    ws.conditional_formatting.add("G5:G19", CellIsRule(operator="greaterThanOrEqual", formula=["100000"], fill=PatternFill("solid", fgColor=WARN_FILL)))

    summary = wb.create_sheet("월간요약")
    style_sheet(summary, "월간 요약", "가계부에서 가장 자주 확인하는 숫자입니다.", 8)
    summary.append([])
    summary.append(["항목", "값", "사용 기능"])
    summary_rows = [
        ["총수입", "=SUM(가계부!F5:F19)", "SUM"],
        ["총지출", "=SUM(가계부!G5:G19)", "SUM"],
        ["남은돈", "=B5-B6", "수입-지출"],
        ["지출건수", '=COUNTIF(가계부!B5:B19,"지출")', "COUNTIF"],
        ["하루 평균 지출", "=B6/30", "나누기"],
        ["저축 가능 여부", '=IF(B7>=300000,"가능","조정필요")', "IF"],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.append([])
    summary.append(["결제수단", "지출합계", "건수"])
    for r, method in enumerate(["카드", "현금", "계좌이체"], start=13):
        summary.append([method, f'=SUMIF(가계부!E5:E19,A{r},가계부!G5:G19)', f'=COUNTIF(가계부!E5:E19,A{r})'])
    style_table(summary, 4, 10, 1, 3)
    style_table(summary, 12, 15, 1, 3)
    set_widths(summary, [16, 14, 14, 4, 12, 12, 12, 12])
    for r in list(range(5, 10)) + list(range(13, 16)):
        summary[f"B{r}"].number_format = '#,##0"원"'

    cats = wb.create_sheet("항목별예산")
    style_sheet(cats, "항목별 예산 확인", "SUMIF와 IF로 예산 초과 항목을 찾습니다.", 6)
    cats.append([])
    cats.append(["항목", "예산", "실제지출", "차이", "판정", "메모"])
    budget_rows = [
        ["식비", 230000],
        ["주거", 650000],
        ["교통", 70000],
        ["통신", 70000],
        ["의료", 50000],
        ["생활용품", 60000],
        ["보험", 150000],
        ["문화", 50000],
        ["경조사", 80000],
        ["기타", 50000],
    ]
    for r, row in enumerate(budget_rows, start=5):
        cats.append([row[0], row[1], f'=SUMIF(가계부!C5:C19,A{r},가계부!G5:G19)', f"=B{r}-C{r}", f'=IF(C{r}>B{r},"예산초과","정상")', ""])
    style_table(cats, 4, 14, 1, 6)
    cats.auto_filter.ref = "A4:F14"
    cats.freeze_panes = "A5"
    set_widths(cats, [12, 12, 12, 12, 12, 16])
    for r in range(5, 15):
        for col in "BCD":
            cats[f"{col}{r}"].number_format = '#,##0"원"'
    cats.conditional_formatting.add("E5:E14", CellIsRule(operator="equal", formula=['"예산초과"'], fill=PatternFill("solid", fgColor=WARN_FILL)))

    guide = wb.create_sheet("실습순서")
    style_sheet(guide, "실습 순서", "수업 시간에 이 순서대로 진행합니다.", 3)
    guide.append([])
    guide.append(["순서", "할 일", "확인"])
    guide_rows = [
        [1, "가계부 시트에서 수입과 지출 입력 구조를 확인합니다.", ""],
        [2, "월간요약 시트에서 총수입, 총지출, 남은돈 수식을 확인합니다.", ""],
        [3, "항목별예산 시트에서 식비, 주거, 교통비 합계를 확인합니다.", ""],
        [4, "예산초과 항목을 조건부 서식 색으로 찾습니다.", ""],
        [5, "가계부 시트 필터로 카드 지출만 확인합니다.", ""],
        [6, "가계부 시트에서 지출액 큰 순서로 정렬합니다.", ""],
    ]
    for row in guide_rows:
        guide.append(row)
    style_table(guide, 4, 10, 1, 3)
    set_widths(guide, [8, 44, 12])

    save(wb, FINAL_DIR / "최종실습_우리집_월간가계부.xlsx")


def create_inventory_project() -> None:
    md = """# 최종 실습 프로젝트: 재고 관리표

## 실습 목표
작은 매장, 사무실, 창고에서 자주 쓰는 재고 관리표를 완성합니다.
품목별 입고, 출고, 현재재고, 안전재고를 정리하고 발주가 필요한 품목을 찾습니다.

## 실제로 자주 쓰는 기능
- SUM으로 전체 재고금액을 계산합니다.
- SUMIF로 품목별 입고수량과 출고수량을 계산합니다.
- IF로 발주필요 여부를 표시합니다.
- COUNTIF로 발주필요 품목 수를 셉니다.
- MIN으로 가장 적게 남은 재고를 확인합니다.
- 정렬로 재고 부족 품목을 위로 올립니다.
- 필터로 발주필요 품목만 봅니다.
- 조건부 서식으로 부족한 재고를 자동 표시합니다.

## 완성 후 확인 질문
- 현재 발주가 필요한 품목은 몇 개인가요?
- 가장 먼저 발주해야 할 품목은 무엇인가요?
- 현재 재고금액이 가장 큰 품목은 무엇인가요?
- 이번 달 입고와 출고가 많은 품목은 무엇인가요?

## 사용 파일
`최종실습_재고관리표.xlsx`
"""
    (FINAL_DIR / "최종실습_재고관리표.md").write_text(md, encoding="utf-8")

    wb = Workbook()

    items = wb.active
    items.title = "재고현황"
    style_sheet(items, "재고 관리표", "현재 재고와 안전재고를 비교해 발주 필요 품목을 찾습니다.", 10)
    items.append([])
    items.append(["품목코드", "품목명", "분류", "기초재고", "입고수량", "출고수량", "현재재고", "안전재고", "단가", "재고금액", "상태"])
    rows = [
        ["P-001", "원두 1kg", "원재료", 12, 20, 26, 10, 18500],
        ["P-002", "우유 1L", "원재료", 30, 50, 58, 15, 2500],
        ["P-003", "종이컵", "소모품", 600, 1000, 1280, 300, 55],
        ["P-004", "포장봉투", "소모품", 450, 500, 720, 250, 80],
        ["P-005", "샌드위치 포장지", "소모품", 180, 300, 410, 120, 120],
        ["P-006", "머그컵", "상품", 24, 20, 31, 10, 5200],
        ["P-007", "원두 200g", "상품", 18, 30, 39, 12, 7200],
        ["P-008", "영수증 용지", "소모품", 8, 10, 14, 8, 960],
    ]
    for r, row in enumerate(rows, start=5):
        items.append(row[:4] + [row[4], row[5], f"=D{r}+E{r}-F{r}", row[6], row[7], f"=G{r}*I{r}", f'=IF(G{r}<H{r},"발주필요","정상")'])
    style_table(items, 4, 12, 1, 11)
    items.auto_filter.ref = "A4:K12"
    items.freeze_panes = "A5"
    set_widths(items, [10, 18, 10, 10, 10, 10, 10, 10, 10, 12, 12])
    for r in range(5, 13):
        items[f"I{r}"].number_format = '#,##0"원"'
        items[f"J{r}"].number_format = '#,##0"원"'
    items.conditional_formatting.add("K5:K12", CellIsRule(operator="equal", formula=['"발주필요"'], fill=PatternFill("solid", fgColor=WARN_FILL)))
    items.conditional_formatting.add("G5:G12", FormulaRule(formula=["$G5<$H5"], fill=PatternFill("solid", fgColor=WARN_FILL)))

    log = wb.create_sheet("입출고기록")
    style_sheet(log, "입출고 기록", "입고와 출고 내역을 날짜순으로 기록합니다.", 8)
    log.append([])
    log.append(["날짜", "구분", "품목코드", "품목명", "수량", "담당자", "거래처/사용처", "메모"])
    log_rows = [
        [date(2026, 6, 1), "입고", "P-001", "원두 1kg", 20, "김지훈", "서울푸드", ""],
        [date(2026, 6, 1), "출고", "P-003", "종이컵", 400, "박선영", "매장사용", ""],
        [date(2026, 6, 2), "입고", "P-002", "우유 1L", 50, "김지훈", "삼익식품", ""],
        [date(2026, 6, 3), "출고", "P-001", "원두 1kg", 10, "이민호", "매장사용", ""],
        [date(2026, 6, 4), "출고", "P-004", "포장봉투", 240, "박선영", "포장사용", ""],
        [date(2026, 6, 5), "입고", "P-006", "머그컵", 20, "김지훈", "에이스비품", ""],
        [date(2026, 6, 8), "출고", "P-007", "원두 200g", 18, "이민호", "판매", ""],
        [date(2026, 6, 9), "출고", "P-008", "영수증 용지", 6, "박선영", "매장사용", ""],
    ]
    for row in log_rows:
        log.append(row)
    style_table(log, 4, 12, 1, 8)
    log.auto_filter.ref = "A4:H12"
    log.freeze_panes = "A5"
    set_widths(log, [12, 8, 10, 18, 8, 10, 16, 16])
    for r in range(5, 13):
        log[f"A{r}"].number_format = "yyyy-mm-dd"

    summary = wb.create_sheet("요약")
    style_sheet(summary, "재고 요약", "발주 대상과 재고금액을 한눈에 확인합니다.", 7)
    summary.append([])
    summary.append(["항목", "값", "사용 기능"])
    summary_rows = [
        ["전체 재고금액", "=SUM(재고현황!J5:J12)", "SUM"],
        ["발주필요 품목 수", '=COUNTIF(재고현황!K5:K12,"발주필요")', "COUNTIF"],
        ["가장 적은 현재재고", "=MIN(재고현황!G5:G12)", "MIN"],
        ["전체 입고수량", "=SUM(재고현황!E5:E12)", "SUM"],
        ["전체 출고수량", "=SUM(재고현황!F5:F12)", "SUM"],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.append([])
    summary.append(["분류", "재고금액", "품목수"])
    for r, category in enumerate(["원재료", "소모품", "상품"], start=12):
        summary.append([category, f'=SUMIF(재고현황!C5:C12,A{r},재고현황!J5:J12)', f'=COUNTIF(재고현황!C5:C12,A{r})'])
    style_table(summary, 4, 9, 1, 3)
    style_table(summary, 11, 14, 1, 3)
    set_widths(summary, [18, 14, 14, 4, 16, 12, 12])
    summary["B5"].number_format = '#,##0"원"'
    for r in range(12, 15):
        summary[f"B{r}"].number_format = '#,##0"원"'

    order = wb.create_sheet("발주목록")
    style_sheet(order, "발주 목록", "발주필요 품목을 옮겨 적고 주문 수량을 결정합니다.", 8)
    order.append([])
    order.append(["품목코드", "품목명", "현재재고", "안전재고", "부족수량", "권장발주", "거래처", "처리"])
    order_rows = [
        ["P-001", "원두 1kg", "=재고현황!G5", "=재고현황!H5"],
        ["P-004", "포장봉투", "=재고현황!G8", "=재고현황!H8"],
        ["P-005", "샌드위치 포장지", "=재고현황!G9", "=재고현황!H9"],
        ["P-008", "영수증 용지", "=재고현황!G12", "=재고현황!H12"],
    ]
    vendors = ["서울푸드", "정다운상사", "에이스비품", "에이스비품"]
    for r, row in enumerate(order_rows, start=5):
        order.append(row + [f"=D{r}-C{r}", f'=IF(E{r}>0,E{r}+10,0)', vendors[r - 5], ""])
    style_table(order, 4, 8, 1, 8)
    order.auto_filter.ref = "A4:H8"
    order.freeze_panes = "A5"
    set_widths(order, [10, 18, 10, 10, 10, 10, 14, 12])
    order.conditional_formatting.add("E5:E8", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=WARN_FILL)))

    guide = wb.create_sheet("실습순서")
    style_sheet(guide, "실습 순서", "수업 시간에 이 순서대로 진행합니다.", 3)
    guide.append([])
    guide.append(["순서", "할 일", "확인"])
    guide_rows = [
        [1, "재고현황 시트에서 현재재고 수식을 확인합니다.", ""],
        [2, "상태 열의 IF 수식으로 발주필요 여부를 확인합니다.", ""],
        [3, "필터로 발주필요 품목만 봅니다.", ""],
        [4, "현재재고가 적은 순서로 정렬합니다.", ""],
        [5, "요약 시트에서 전체 재고금액과 발주필요 품목 수를 확인합니다.", ""],
        [6, "발주목록 시트에서 권장발주 수량을 확인합니다.", ""],
    ]
    for row in guide_rows:
        guide.append(row)
    style_table(guide, 4, 10, 1, 3)
    set_widths(guide, [8, 48, 12])

    save(wb, FINAL_DIR / "최종실습_재고관리표.xlsx")


def main() -> None:
    clean_dirs()
    for item in ITEMS:
        write_md(item)
        workbook_for_item(item)
    create_final_project()
    create_household_budget_project()
    create_inventory_project()
    print(f"교재: {DOC_DIR}")
    print(f"항목별 실습예제: {EX_DIR}")
    print(f"최종실습: {FINAL_DIR}")


if __name__ == "__main__":
    main()
